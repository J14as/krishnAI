from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Learning Progress"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 30

# Title
ws['A1'] = "KrishnAI Project - Learning Progress Tracker"
ws['A1'].font = Font(bold=True, size=14, color="4472C4")
ws.merge_cells('A1:E1')
ws['A1'].alignment = center_align

# Subtitle with date
ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
ws['A2'].font = Font(italic=True, size=10)
ws.merge_cells('A2:E2')

# Headers
headers = ["Day", "Date", "Feature/Topic Covered", "Technologies Used", "Learning Outcomes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Data rows
progress_data = [
    [1, "Day 1", "Project Setup & Backend Architecture", 
     "Flask, SQLAlchemy, Flask-Login", 
     "Understood Flask application structure, database modeling, user authentication system"],
    
    [2, "Day 2", "User Authentication System", 
     "Werkzeug Security, Password Hashing, Flask-Login", 
     "Implemented secure user signup/login, password hashing, session management"],
    
    [3, "Day 3", "AI Integration - Groq API", 
     "Groq API, LLaMA 3.1 Model, Environment Variables, API Integration", 
     "Integrated AI chatbot with Groq API, learned how to use external AI models, prompt engineering"],
    
    [4, "Day 4", "Chat Feature Implementation", 
     "Flask Routes, Database Operations, Form Handling, Jinja2 Templates", 
     "Built interactive chat feature, stored conversations in database, rendered responses dynamically"],
    
    [5, "Day 5", "Daily Shloka Feature", 
     "Shloka Service, Data Management", 
     "Implemented daily wisdom quotes feature, learned data retrieval and presentation"],
    
    [6, "Day 6", "Reflection Module", 
     "CRUD Operations, Flask-SQLAlchemy, Delete Operations", 
     "Built reflection journal feature, implemented full CRUD operations, security checks for data ownership"],
    
    [7, "Day 7", "Dashboard & Navigation", 
     "Frontend Design, HTML/CSS, User Interface", 
     "Created user dashboard, implemented navigation system, improved UI/UX"],
    
    [8, "Day 8", "Frontend Development", 
     "HTML5, CSS3, JavaScript, Theme Management", 
     "Designed responsive templates, implemented theme switching, enhanced user experience"],
    
    [9, "Day 9", "Database Schema & Models", 
     "SQLAlchemy ORM, Database Design, Relationships", 
     "Designed three main models (User, Chat, Reflection), understood ORM concepts, data relationships"],
    
    [10, "Day 10", "Authentication & Security", 
     "Login Required Decorators, User Verification, Security Checks", 
     "Implemented protected routes, user verification, security best practices"],
]

row = 5
for data in progress_data:
    day_num, day_label, feature, tech, outcomes = data
    
    ws.cell(row=row, column=1).value = day_num
    ws.cell(row=row, column=2).value = day_label
    ws.cell(row=row, column=3).value = feature
    ws.cell(row=row, column=4).value = tech
    ws.cell(row=row, column=5).value = outcomes
    
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        if col == 1:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
    
    row += 1

# Summary section
summary_row = row + 2
ws[f'A{summary_row}'] = "PROJECT SUMMARY"
ws[f'A{summary_row}'].font = subheader_font
ws[f'A{summary_row}'].fill = subheader_fill
ws.merge_cells(f'A{summary_row}:E{summary_row}')

summary_content = [
    ["Project Name:", "KrishnAI - An AI-Powered Krishna Wisdom Chatbot"],
    ["Project Type:", "Full-Stack Web Application"],
    ["Tech Stack:", "Python, Flask, SQLAlchemy, Groq AI API, HTML/CSS/JavaScript"],
    ["Key Features:", "User Authentication, AI Chat, Daily Shloka, Reflection Journal, Dashboard"],
    ["Database:", "SQLite with SQLAlchemy ORM"],
    ["Deployment Ready:", "Yes - Can be deployed on any WSGI server"],
]

summary_row += 1
for item in summary_content:
    ws[f'A{summary_row}'] = item[0]
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'] = item[1]
    ws.merge_cells(f'B{summary_row}:E{summary_row}')
    for col in range(1, 6):
        ws.cell(row=summary_row, column=col).border = border
    summary_row += 1

# Skills Acquired section
skills_row = summary_row + 2
ws[f'A{skills_row}'] = "SKILLS ACQUIRED"
ws[f'A{skills_row}'].font = subheader_font
ws[f'A{skills_row}'].fill = subheader_fill
ws.merge_cells(f'A{skills_row}:E{skills_row}')

skills = [
    "✓ Full-Stack Web Development",
    "✓ Python Programming & Flask Framework",
    "✓ Database Design & SQLAlchemy ORM",
    "✓ User Authentication & Security",
    "✓ API Integration (Groq AI)",
    "✓ CRUD Operations",
    "✓ Frontend Development (HTML/CSS/JavaScript)",
    "✓ Version Control & Git",
    "✓ Environment Variable Management",
    "✓ MVC Architecture Pattern",
]

skills_row += 1
skills_col = 1
for i, skill in enumerate(skills):
    if i % 2 == 0:
        ws[f'A{skills_row}'] = skill
        ws.merge_cells(f'A{skills_row}:B{skills_row}')
    else:
        ws[f'D{skills_row}'] = skill
        ws.merge_cells(f'D{skills_row}:E{skills_row}')
        skills_row += 1
    
    for col in range(1, 6):
        ws.cell(row=skills_row, column=col).border = border

# Save the workbook
output_path = "KrishnAI_Learning_Progress.xlsx"
wb.save(output_path)
print(f"✓ Excel sheet created successfully: {output_path}")
